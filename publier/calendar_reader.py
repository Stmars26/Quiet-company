"""Read the QC Master Social Calendar and resolve image assets."""

from __future__ import annotations

import os
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pytz

logger = logging.getLogger(__name__)

CET = pytz.timezone("Europe/Amsterdam")

# Calendar columns (1-indexed)
COL_WEEK = 1
COL_DATE = 2
COL_DAY = 3
COL_THEME = 4
COL_PLATFORM = 5
COL_CONTENT_TYPE = 6
COL_CAPTION = 7
COL_HASHTAGS = 8
COL_VISUAL_ASSET = 9
COL_CTA = 10
COL_STATUS = 11
COL_NOTES = 12

# Map calendar asset names → actual filenames in static/images/launch/
ASSET_MAP = {
    "character-spotlight-claire.png": "post-claire-spotlight.png",
    "character-spotlight-elena.png": "post-elena-spotlight.png",
    "character-spotlight-marcus.png": "post-marcus-spotlight.png",
    "character-spotlight-tess.png": "post-tess-spotlight.png",
    "eight-rooms-overview.png": "post-eight-rooms.png",
    "launch-offer-card.png": "post-launch-offer.png",
    "brand-philosophy-carousel.png": "post-philosophy.png",
    "quote-card-gap.png": "quote-2am.png",
    "quote-card-mindful.png": "quote-consistency.png",
    "og-image.png": "og-image.png",
}

SUPPORTED_PLATFORMS = {"Instagram", "X (Twitter)"}

# Year to use when parsing "Mar 31" style dates
CALENDAR_YEAR = 2026


def parse_calendar_date(date_str: str) -> datetime | None:
    """Parse 'Mar 31' into a CET-aware datetime."""
    if not date_str or not isinstance(date_str, str):
        return None
    try:
        dt = datetime.strptime(f"{date_str} {CALENDAR_YEAR}", "%b %d %Y")
        return CET.localize(dt)
    except ValueError:
        logger.warning(f"Could not parse date: {date_str}")
        return None


def resolve_asset(asset_name: str, launch_dir: Path) -> Path | None:
    """Resolve a calendar asset name to an actual file path."""
    if not asset_name or asset_name == "None":
        return None

    # Direct match in launch dir
    direct = launch_dir / asset_name
    if direct.exists():
        return direct

    # Try mapped name
    mapped = ASSET_MAP.get(asset_name)
    if mapped:
        mapped_path = launch_dir / mapped
        if mapped_path.exists():
            return mapped_path

    # Try parent images dir
    images_dir = launch_dir.parent
    direct_parent = images_dir / asset_name
    if direct_parent.exists():
        return direct_parent

    return None


def read_calendar(
    xlsx_path: str | Path,
    launch_dir: str | Path,
    platforms: set[str] | None = None,
) -> list[dict]:
    """
    Read the Master Calendar sheet and return a list of post dicts.

    Each dict has:
        row: int (1-indexed spreadsheet row, used as unique ID)
        week: str
        date: datetime (CET)
        day: str
        theme: str
        platform: str
        content_type: str
        caption: str
        hashtags: str
        asset_path: Path | None
        asset_name: str | None
        cta: str
        status: str
        notes: str
    """
    platforms = platforms or SUPPORTED_PLATFORMS
    xlsx_path = Path(xlsx_path)
    launch_dir = Path(launch_dir)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Calendar not found: {xlsx_path}")

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb["Master Calendar"]

    posts = []
    for row_num in range(2, ws.max_row + 1):
        platform = ws.cell(row_num, COL_PLATFORM).value
        if not platform or platform not in platforms:
            continue

        date_str = ws.cell(row_num, COL_DATE).value
        dt = parse_calendar_date(str(date_str) if date_str else "")
        if not dt:
            continue

        caption = ws.cell(row_num, COL_CAPTION).value or ""
        hashtags = ws.cell(row_num, COL_HASHTAGS).value or ""
        asset_name = ws.cell(row_num, COL_VISUAL_ASSET).value
        asset_path = resolve_asset(asset_name, launch_dir) if asset_name else None

        # Combine caption + hashtags for posting
        full_caption = caption.strip()
        if hashtags and hashtags != "None":
            full_caption = f"{full_caption}\n\n{hashtags.strip()}"

        posts.append({
            "row": row_num,
            "week": ws.cell(row_num, COL_WEEK).value or "",
            "date": dt,
            "day": ws.cell(row_num, COL_DAY).value or "",
            "theme": ws.cell(row_num, COL_THEME).value or "",
            "platform": platform,
            "content_type": ws.cell(row_num, COL_CONTENT_TYPE).value or "",
            "caption": full_caption,
            "hashtags": hashtags,
            "asset_path": asset_path,
            "asset_name": asset_name,
            "cta": ws.cell(row_num, COL_CTA).value or "",
            "status": ws.cell(row_num, COL_STATUS).value or "",
            "notes": ws.cell(row_num, COL_NOTES).value or "",
        })

    wb.close()
    logger.info(f"Loaded {len(posts)} posts for {platforms} from {xlsx_path.name}")
    return posts


def get_posts_for_date(posts: list[dict], target_date: datetime) -> list[dict]:
    """Filter posts scheduled for a specific date."""
    target = target_date.date()
    return [p for p in posts if p["date"].date() == target]
